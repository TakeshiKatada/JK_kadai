#! python3
# readCensusExcel.py - 郡ごとに人口と人口調査標準地域の数を集計する

import openpyxl, pprint #1
print('ワークブックを開いています...')
wb = openpyxl.load_workbook('censuspopdata.xlsx') #2
sheet = wb.get_sheet_by_name('Population by Census Tract') #3
county_data = {}

# TODO: county_dataに郡の人口と地域数を格納する
print('行を読み込んでます...')
for row in range(2, sheet.max_row + 1): #4
    # スプレッドシートの1行に、一つの人口調査標準地域データがある
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value

# TODO: 新しいテキストファイルを開き、county_dataの内容を書き込む

for row in range(2, sheet.max_row +1):
    #スプレットシートの1行に、一つの人口調査標準地域のデータがある
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value

    #この州のキーが確実に存在するようにする
    county_data.setdefault(state, {}) #1
    #この州のこの郡のキーが確実に存在するようにする
    county_data[state].setdefault(county, {'tracts': 0, 'pop': 0}) #2

    #各行が人口調査標準地域を表すので、数を1つ増やす
    county_data[state][county]['tracts'] += 1 #3
    #この人口調査標準地域の人口だけ郡の人口を増やす
    county_data[state][county]['pop'] += int(pop) #4

# TODO: 新しいテキストファイルを開き、county_dataの内容を書き込む

print('結果を書き込み中...')
result_file = open('census2010.txt', 'w')
result_file.write('all_data = ' + pprint.pformat(county_data))
result_file.close()
print('完了')
